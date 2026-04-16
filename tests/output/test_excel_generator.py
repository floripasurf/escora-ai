"""Smoke tests for Excel report generation."""

import pytest
import os
from pathlib import Path
from openpyxl import load_workbook
from src.output.excel_generator import generate_excel
from src.output.report_data import (
    ReportData, ReportMetadata, SummaryData,
    BeamRow, SlabRow, BomRow, build_report_data,
)
from src.models.calculation_models import CalculationResult


def _empty_report():
    return ReportData(
        project_name="TEST",
        date="2026-03-25",
        summary=SummaryData(
            total_shores=0, total_load_kn=0.0,
            pe_direito_m=2.80, pe_direito_is_default=False,
            slab_thickness_m=0.12, thickness_is_default=True,
            beam_count=0, slab_count=0, is_valid=True,
        ),
    )


def _report_with_data():
    return ReportData(
        project_name="TEST",
        date="2026-03-25",
        summary=SummaryData(
            total_shores=10, total_load_kn=500.0,
            pe_direito_m=2.80, pe_direito_is_default=True,
            slab_thickness_m=0.12, thickness_is_default=True,
            beam_count=2, slab_count=1, is_valid=True,
        ),
        beam_rows=[
            BeamRow(name="V1", section="14x40 cm", section_width_m=0.14,
                    section_height_m=0.40, length_m=8.0, load_kn_m=12.5,
                    shore_count=7, spacing_m=1.2, shore_model="Escora T-01",
                    is_cantilever=False),
        ],
        slab_rows=[
            SlabRow(panel_id=1, area_m2=20.0, thickness_m=0.12,
                    total_load_kn=100.0, grid="3x3", spacing_x_m=1.5,
                    spacing_y_m=1.2, shore_model="Escora T-02",
                    is_cantilever=False),
        ],
        bom_rows=[
            BomRow(id="ESC-01", model="Escora T-01", manufacturer="Generico",
                   quantity=7, capacity_kn=20.0, height_min_m=1.80,
                   height_max_m=3.20, weight_kg=11.0, total_weight_kg=77.0,
                   price_brl=65.0, total_price_brl=455.0),
        ],
        warnings=["Pe-direito padrao"],
    )


class TestExcelGenerator:
    def test_generates_file(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        generate_excel(_report_with_data(), path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

    def test_has_three_sheets(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        generate_excel(_report_with_data(), path)
        wb = load_workbook(path)
        assert wb.sheetnames == ["BOM", "Vigas", "Lajes", "Volumes", "Consumo"]

    def test_bom_sheet_headers(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        generate_excel(_report_with_data(), path)
        wb = load_workbook(path)
        ws = wb["BOM"]
        headers = [cell.value for cell in ws[1]]
        assert "id" in headers
        assert "modelo" in headers
        assert "quantidade" in headers

    def test_bom_sheet_data(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        generate_excel(_report_with_data(), path)
        wb = load_workbook(path)
        ws = wb["BOM"]
        assert ws.cell(row=2, column=1).value == "ESC-01"
        assert ws.cell(row=2, column=4).value == 7

    def test_vigas_sheet_data(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        generate_excel(_report_with_data(), path)
        wb = load_workbook(path)
        ws = wb["Vigas"]
        assert ws.cell(row=2, column=1).value == "V1"

    def test_lajes_sheet_data(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        generate_excel(_report_with_data(), path)
        wb = load_workbook(path)
        ws = wb["Lajes"]
        assert ws.cell(row=2, column=1).value == 1  # panel_id

    def test_empty_report_generates(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        generate_excel(_empty_report(), path)
        assert os.path.exists(path)
        wb = load_workbook(path)
        assert wb.sheetnames == ["BOM", "Vigas", "Lajes", "Volumes", "Consumo"]
