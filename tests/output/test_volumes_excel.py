"""Testes da aba `Volumes` no Excel e do breakdown no ReportData."""


from openpyxl import load_workbook

from src.models.calculation_models import (
    CalculationResult, VolumeBreakdownEntry,
)
from src.output.excel_generator import generate_excel
from src.output.report_data import (
    ReportMetadata, build_report_data,
)


def _sample_calc():
    return CalculationResult(
        beam_results=[],
        slab_results=[],
        pe_direito_m=2.80,
        slab_volume_gross_m3=256.79,
        beam_volume_deducted_m3=1.80,
        pillar_volume_deducted_m3=0.90,
        total_volume_m3=254.09,
        volume_breakdown=[
            VolumeBreakdownEntry(
                category="laje", label="Laje 1 (Quarto 1)",
                area_m2=80.0, pe_direito_m=2.80, volume_m3=224.0,
                centroid_x=5.0, centroid_y=5.0,
            ),
            VolumeBreakdownEntry(
                category="beiral", label="Beiral 1",
                area_m2=6.0, pe_direito_m=2.80, volume_m3=16.80,
                centroid_x=0.5, centroid_y=5.0,
            ),
            VolumeBreakdownEntry(
                category="platibanda", label="Platibanda 1",
                area_m2=5.71, pe_direito_m=2.80, volume_m3=15.99,
                centroid_x=10.0, centroid_y=5.0,
            ),
        ],
    )


def _sample_report():
    meta = ReportMetadata(
        project_name="TEST", date="2026-04-16",
        scale=1.0, dxf_filename="test.dxf",
    )
    return build_report_data(_sample_calc(), meta)


class TestReportDataVolume:
    def test_volume_rows_populated(self):
        report = _sample_report()
        assert len(report.volume_rows) == 3
        assert report.volume_rows[0].element == "Laje 1 (Quarto 1)"
        assert report.volume_rows[1].category == "beiral"
        assert report.volume_rows[2].category_label == "Platibanda"

    def test_volume_totals_populated(self):
        report = _sample_report()
        assert report.volume_totals["bruto_m3"] == 256.79
        assert report.volume_totals["vigas_m3"] == 1.80
        assert report.volume_totals["pilares_m3"] == 0.90
        assert report.volume_totals["liquido_m3"] == 254.09


class TestExcelVolumesTab:
    def test_volumes_sheet_exists(self, tmp_path):
        path = str(tmp_path / "vol.xlsx")
        generate_excel(_sample_report(), path)
        wb = load_workbook(path)
        assert "Volumes" in wb.sheetnames

    def test_volumes_headers(self, tmp_path):
        path = str(tmp_path / "vol.xlsx")
        generate_excel(_sample_report(), path)
        wb = load_workbook(path)
        ws = wb["Volumes"]
        headers = [cell.value for cell in ws[1]]
        assert "categoria" in headers
        assert "elemento" in headers
        assert "area_m2" in headers
        assert "pe_direito_m" in headers
        assert "volume_m3" in headers

    def test_volumes_row_data(self, tmp_path):
        path = str(tmp_path / "vol.xlsx")
        generate_excel(_sample_report(), path)
        wb = load_workbook(path)
        ws = wb["Volumes"]
        # Linha 2 = primeiro elemento
        assert ws.cell(row=2, column=1).value == "Laje"
        assert ws.cell(row=2, column=2).value == "Laje 1 (Quarto 1)"
        assert ws.cell(row=2, column=5).value == 224.0

    def test_volumes_totals_rows(self, tmp_path):
        path = str(tmp_path / "vol.xlsx")
        generate_excel(_sample_report(), path)
        wb = load_workbook(path)
        ws = wb["Volumes"]
        # 3 elementos + 1 linha em branco + 4 linhas de total = linha 9 base.
        # Procurar célula com "TOTAL BRUTO" na coluna 2.
        found = False
        for row in range(2, 15):
            val = ws.cell(row=row, column=2).value
            if val == "TOTAL BRUTO":
                assert ws.cell(row=row, column=5).value == 256.79
                found = True
            if val == "LÍQUIDO":
                assert ws.cell(row=row, column=5).value == 254.09
        assert found, "Totalizador não encontrado na aba Volumes"

    def test_empty_breakdown_still_writes_totals(self, tmp_path):
        meta = ReportMetadata(
            project_name="T", date="2026-04-16",
            scale=1.0, dxf_filename="t.dxf",
        )
        empty_calc = CalculationResult(
            beam_results=[], slab_results=[],
            pe_direito_m=2.80,
            slab_volume_gross_m3=0.0, beam_volume_deducted_m3=0.0,
            pillar_volume_deducted_m3=0.0, total_volume_m3=0.0,
            volume_breakdown=[],
        )
        report = build_report_data(empty_calc, meta)
        path = str(tmp_path / "empty_vol.xlsx")
        generate_excel(report, path)
        wb = load_workbook(path)
        ws = wb["Volumes"]
        # Totais ainda devem aparecer (todos zero)
        any_total = False
        for row in range(2, 10):
            if ws.cell(row=row, column=2).value in ("TOTAL BRUTO", "LÍQUIDO"):
                any_total = True
        assert any_total
