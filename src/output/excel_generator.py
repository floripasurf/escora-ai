"""Generate Excel workbook with BOM, Vigas, and Lajes sheets."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from src.output.report_data import ReportData


BOM_HEADERS = [
    "id", "modelo", "fabricante", "quantidade", "capacidade_kn",
    "altura_min_m", "altura_max_m", "peso_unitario_kg", "peso_total_kg",
    "preco_unitario_brl", "preco_total_brl",
]

VIGAS_HEADERS = [
    "nome", "secao_largura_m", "secao_altura_m", "comprimento_m",
    "carga_linear_kn_m", "qtd_escoras", "espacamento_m",
    "modelo_escora", "balanco",
]

LAJES_HEADERS = [
    "painel", "area_m2", "espessura_m", "carga_total_kn",
    "grid_nx", "grid_ny", "espacamento_x_m", "espacamento_y_m",
    "modelo_escora", "balanco",
]

VOLUMES_HEADERS = [
    "categoria", "elemento", "area_m2", "pe_direito_m", "volume_m3",
]

CONSUMO_HEADERS = [
    "pe_direito_m", "area_m2", "volume_bruto_m3", "volume_liquido_m3",
    "escoras_kg", "acessorios_kg", "total_kg",
    "taxa_kg_m3_bruto", "taxa_kg_m3_liquido",
]


def _write_sheet(ws, headers, rows):
    """Write headers + data rows to a worksheet."""
    bold = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold
    for row_idx, row_data in enumerate(rows, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, len(rows) + 2):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 30)


def generate_excel(report: ReportData, output_path: str) -> str:
    """Generate Excel workbook from ReportData.

    Args:
        report: Normalized report data.
        output_path: Path to write .xlsx file.

    Returns:
        The output_path written to.
    """
    wb = Workbook()

    # Sheet 1: BOM
    ws_bom = wb.active
    ws_bom.title = "BOM"
    bom_rows = [
        [r.id, r.model, r.manufacturer, r.quantity, r.capacity_kn,
         r.height_min_m, r.height_max_m, r.weight_kg, r.total_weight_kg,
         r.price_brl, r.total_price_brl]
        for r in report.bom_rows
    ]
    _write_sheet(ws_bom, BOM_HEADERS, bom_rows)

    # Sheet 2: Vigas
    ws_vigas = wb.create_sheet("Vigas")
    vigas_rows = [
        [r.name, r.section_width_m, r.section_height_m, r.length_m,
         r.load_kn_m, r.shore_count, r.spacing_m, r.shore_model,
         "Sim" if r.is_cantilever else "Não"]
        for r in report.beam_rows
    ]
    _write_sheet(ws_vigas, VIGAS_HEADERS, vigas_rows)

    # Sheet 3: Lajes
    ws_lajes = wb.create_sheet("Lajes")
    lajes_rows = []
    for r in report.slab_rows:
        parts = r.grid.split("x")
        nx = int(parts[0]) if len(parts) == 2 else 0
        ny = int(parts[1]) if len(parts) == 2 else 0
        lajes_rows.append([
            r.panel_id, r.area_m2, r.thickness_m, r.total_load_kn,
            nx, ny, r.spacing_x_m, r.spacing_y_m, r.shore_model,
            "Sim" if r.is_cantilever else "Não",
        ])
    _write_sheet(ws_lajes, LAJES_HEADERS, lajes_rows)

    # Sheet 4: Volumes (breakdown didático por painel/elemento)
    _write_volumes_sheet(wb, report)

    # Sheet 5: Consumo (resumo agregado por pé-direito — orçamento interno)
    _write_consumption_sheet(wb, report)

    wb.save(output_path)
    return output_path


def _write_volumes_sheet(wb, report) -> None:
    """Escreve aba `Volumes` com breakdown por elemento + 4 linhas de totais.

    Layout:
        | categoria | elemento            | area_m2 | pe_direito_m | volume_m3 |
        | Laje      | Laje 1 (Quarto 1)   |  60.50  | 2.80         | 169.40    |
        | Beiral    | Beiral 1            |   4.30  | 2.80         |  12.04    |
        | ...                                                                  |
        |                TOTAL BRUTO                              | 240.80    |
        |                (-) Vigas                                |   1.80    |
        |                (-) Pilares                              |   0.90    |
        |                LÍQUIDO (bold)                           | 238.10    |
    """
    ws = wb.create_sheet("Volumes")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    # Cabeçalho
    for col, header in enumerate(VOLUMES_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold

    # Linhas de elementos (ordenadas por categoria, mantém ordem de volume_rows)
    last_row = 1
    for row_idx, r in enumerate(report.volume_rows, 2):
        ws.cell(row=row_idx, column=1, value=r.category_label)
        ws.cell(row=row_idx, column=2, value=r.element)
        ws.cell(row=row_idx, column=3, value=round(r.area_m2, 2))
        ws.cell(row=row_idx, column=4, value=round(r.pe_direito_m, 2))
        ws.cell(row=row_idx, column=5, value=round(r.volume_m3, 2))
        last_row = row_idx

    totals = report.volume_totals or {}
    bruto = float(totals.get("bruto_m3", 0.0))
    vigas = float(totals.get("vigas_m3", 0.0))
    pilares = float(totals.get("pilares_m3", 0.0))
    liquido = float(totals.get("liquido_m3", 0.0))

    # Separador em branco + 4 linhas de totais
    summary_rows = [
        ("TOTAL BRUTO",  bruto,   True,  False),
        ("(-) Vigas",    vigas,   False, False),
        ("(-) Pilares",  pilares, False, False),
        ("LÍQUIDO",      liquido, True,  True),  # negrito destaque
    ]
    for i, (label, value, is_bold, is_highlight) in enumerate(summary_rows, 1):
        row = last_row + 1 + i
        label_cell = ws.cell(row=row, column=2, value=label)
        value_cell = ws.cell(row=row, column=5, value=round(value, 2))
        if is_bold or is_highlight:
            label_cell.font = bold
            value_cell.font = bold
        label_cell.alignment = center

    # Freeze header
    ws.freeze_panes = "A2"
    # Auto-width
    widths = {1: 14, 2: 30, 3: 10, 4: 14, 5: 12}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_consumption_sheet(wb, report) -> None:
    """Aba `Consumo` — resumo agregado por pé-direito (orçamento interno).

    Layout:
        | pe_direito_m | area_m2 | volume_bruto_m3 | volume_liquido_m3 |
        | escoras_kg | acessorios_kg | total_kg |
        | taxa_kg_m3_bruto | taxa_kg_m3_liquido |

    Linha final TOTAL em negrito agregando colunas e taxas ponderadas.
    """
    ws = wb.create_sheet("Consumo")
    bold = Font(bold=True)

    # Cabeçalho
    for col, header in enumerate(CONSUMO_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold

    # Linhas por pé-direito
    last_row = 1
    for row_idx, r in enumerate(report.consumption_rows, 2):
        ws.cell(row=row_idx, column=1, value=round(r.pe_direito_m, 2))
        ws.cell(row=row_idx, column=2, value=round(r.area_m2, 2))
        ws.cell(row=row_idx, column=3, value=round(r.volume_bruto_m3, 2))
        ws.cell(row=row_idx, column=4, value=round(r.volume_liquido_m3, 2))
        ws.cell(row=row_idx, column=5, value=round(r.shores_weight_kg, 2))
        ws.cell(row=row_idx, column=6, value=round(r.accessories_weight_kg, 2))
        ws.cell(row=row_idx, column=7, value=round(r.total_weight_kg, 2))
        ws.cell(row=row_idx, column=8, value=round(r.rate_kg_m3_bruto, 2))
        ws.cell(row=row_idx, column=9, value=round(r.rate_kg_m3_liquido, 2))
        last_row = row_idx

    totals = report.consumption_totals or {}
    if totals:
        total_row = last_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL").font = bold
        ws.cell(row=total_row, column=2, value=round(totals.get("area_m2", 0.0), 2)).font = bold
        ws.cell(row=total_row, column=3, value=round(totals.get("volume_bruto_m3", 0.0), 2)).font = bold
        ws.cell(row=total_row, column=4, value=round(totals.get("volume_liquido_m3", 0.0), 2)).font = bold
        ws.cell(row=total_row, column=5, value=round(totals.get("shores_kg", 0.0), 2)).font = bold
        ws.cell(row=total_row, column=6, value=round(totals.get("accessories_kg", 0.0), 2)).font = bold
        ws.cell(row=total_row, column=7, value=round(totals.get("total_kg", 0.0), 2)).font = bold
        ws.cell(row=total_row, column=8, value=round(totals.get("rate_kg_m3_bruto", 0.0), 2)).font = bold
        ws.cell(row=total_row, column=9, value=round(totals.get("rate_kg_m3_liquido", 0.0), 2)).font = bold

    # Freeze header + larguras
    ws.freeze_panes = "A2"
    widths = {1: 14, 2: 12, 3: 18, 4: 20, 5: 14, 6: 16, 7: 12, 8: 20, 9: 22}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
